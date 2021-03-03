import json
from datetime import timedelta
from pathlib import Path

import os
import zipfile
from django.core.files.base import ContentFile
from django.core.mail import send_mail
from django.core.files.storage import default_storage
from django.template.loader import render_to_string
from django.utils import timezone
from apps.dms.api.category.models import MetaField, Category
from apps.core.rbac.models import User
from apps.dms.documents.elastic_search import Elastic
from apps.dms.documents.models import Documents, DownloadSearchResult
from conf import settings
from conf.celery import app
from openpyxl import Workbook, load_workbook
from conf import licensed
import datetime
import collections


class DMSTasks(object):
    @staticmethod
    def create_zip(user_id, doc_id_list):
        user = User.objects.get(pk=user_id)
        documents = Documents.objects.filter(id__in=doc_id_list).only('filepath', 'filename', 'id')
        print(documents)
        zip_folder_path = os.path.join(settings.MEDIA_ROOT, 'zip_files', str(user.id))

        # create zip folder
        path = Path(zip_folder_path)
        path.mkdir(parents=True, exist_ok=True)

        zip_filename = timezone.now().strftime("%Y_%m_%d_%H_%M_%S")
        media_file = zip_file_path = '{}/{}.zip'.format(zip_folder_path, zip_filename)
        zip_file_path = zipfile.ZipFile(zip_file_path, 'w', zipfile.ZIP_DEFLATED)

        # write file to zip file
        for document in documents:
            path = default_storage.path(document.filepath)
            print(path)
            filename, ext = os.path.splitext(path)
            metadata = json.loads(document.metadata)
            meta_string = ''

            for meta in metadata:
                meta_string += meta.get('value') + '-'

            doc_filename = document.filename.replace(' ', '_')
            name = '{}-{}-{}{}{}'.format(doc_filename, document.doc_type.name, meta_string, document.id, ext)
            zip_file_path.write(path, name)

        zip_file_path.close()
        final_zip_file_path = media_file.replace(settings.MEDIA_ROOT + '/', '')
        instance = DownloadSearchResult.objects.create(**{
            'user': user,
            'path': final_zip_file_path
        })

        # send zip file path as attachment through email
        # subject, from_email, to = 'Your file is ready to download', settings.EMAIL_HOST_USER, user.email
        # download_link = 'http://{}/login/?next={}{}'.format(settings.APP_HOST, settings.MEDIA_URL,
        # final_zip_file_path)
        # html_content = render_to_string('email/download_search_files.html', {
        #     'user_full_name': user.get_full_name(),
        #     'validity_date': (instance.created_at + timedelta(days=7)).strftime(
        #         '%d %B, %Y'),
        #     'download_link': download_link,
        #     'zip_filename': zip_filename
        # })
        # send_mail(subject=subject, message='', from_email=from_email, recipient_list=[to], html_message=html_content,
        #           fail_silently=True)

    # Download all files from DMS search
    @staticmethod
    @app.task(name='download-all-items-from-dms')
    def download_all_items(user, exclude_list, search_params, search_total_loop):
        doc_id_list = []

        for loop in range(0, search_total_loop):
            result = Elastic.search(search_params)

            for data in result.get('data'):
                file_id = int(data.get('id'))

                if file_id not in exclude_list:
                    doc_id_list.append(int(data.get('id')))

            start = search_params.get('start')
            length = search_params.get('length')
            search_params.update({'start': start + length})

        DMSTasks.create_zip(user, doc_id_list)

    # Download selected files from DMS search
    @staticmethod
    @app.task(name='download-selected-items-from-dms')
    def download_selected_items(user_id, doc_id_list):
        DMSTasks.create_zip(user_id, doc_id_list)

    @staticmethod
    @app.task(name='upload_report_export')
    def upload_report_export(doc_type_id, user_id, x_forwarded_for, host, name, to, date_from, date_to):
        print("date-from", date_from)
        print("date-to", date_to)
        date_from = date_from.replace(" ", "+")
        date_to = date_to.replace(" ", "+")
        if date_from is "null" or date_to is "null":
            date_from = None
            date_to = None
        meta_key = []
        edited_value = []
        documents = Documents.objects.filter(doc_type_id=doc_type_id)
        if date_from and date_to:
            documents = documents.filter(uploaded_at__range=[date_from, date_to])
        meta_category = MetaField.objects.filter(doc_id=doc_type_id).order_by('order')
        for c in meta_category:
            meta_key.append(c.title)
        doc_list = []

        for h in documents:
            docTypes = Category.objects.get(id=doc_type_id)
            docTypes = docTypes.get_ancestors(ascending=False, include_self=True)
            docType = '->'.join([str(i) for i in docTypes])
            doc_dict = {}
            user_name = User.objects.get(pk=h.uploader_id).get_full_name()
            doc_dict['uploader'] = user_name
            doc_dict['document_name'] = h.filename
            doc_dict['doc_type_name'] = docType
            doc_list.append(doc_dict)
            metas_element = collections.OrderedDict()
            for key in meta_key:
                metas_element[key] = ""
            metas = json.loads(h.metadata)
            print("metas", metas)
            for key in metas:
                field_key = key['name']
                field_value = key['value']
                metas_element[field_key] = field_value
            edited_value.append(metas_element)
        value_list = ['Uploader', 'Document Name', 'Document Type']
        wb = Workbook()
        ws = wb.active
        for i in range(1, 4):
            ws.cell(row=1, column=i).value = value_list[i - 1]
        col = 4
        for item in meta_key:
            ws.cell(row=1, column=col).value = item
            col = col + 1
        data_col = 1
        for i in range(0, len(doc_list)):
            ws.cell(row=i + 2, column=data_col).value = doc_list[i]['uploader']
            data_col = data_col + 1
            ws.cell(row=i + 2, column=data_col).value = doc_list[i]['doc_type_name']
            data_col = data_col + 1
            ws.cell(row=i + 2, column=data_col).value = doc_list[i]['document_name']
            data_col = data_col + 1
            for k in meta_key:
                ws.cell(row=i + 2, column=data_col).value = edited_value[i][k]
                data_col = data_col + 1
            data_col = 1

        file_name = "Upload Report" + str(datetime.datetime.now()) + '.xlsx'

        if not default_storage.exists('nrb/upload_report'):
            default_storage.save('nrb/upload_report/nrb_report_created.bin', ContentFile('shuvo'))
        wb.save('media/nrb/upload_report/' + file_name)

        subject = "Download Report"
        message = "Your requested file is ready to download"

        if x_forwarded_for:
            address = x_forwarded_for.split(',')[-1].strip()
        else:
            address = host
        link = '{}/api/v1/upload_report_download'.format(address)
        html_content = render_to_string('dms/report/report_mail_template.html',
                                        {
                                            'sent_by': licensed.EMAIL_HOST_USER,
                                            'message': message,
                                            'name': name,
                                            'file_name': file_name,
                                            'file_path': 'nrb/upload_report/' + file_name,
                                            'link': link
                                        }
                                        )
        send_mail(subject=subject, message=message,
                  from_email=licensed.EMAIL_HOST_USER,
                  recipient_list=[to],
                  html_message=html_content,
                  fail_silently=False)